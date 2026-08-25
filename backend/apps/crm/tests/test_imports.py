"""Final-completion-pass: tests for apps/crm/imports.py (real CSV/XLSX
bulk import and export for Leads) and the LeadViewSet actions built on it.
"""
import csv
import io

import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile

from apps.crm.imports import EXPORT_COLUMNS, export_leads_csv, export_leads_xlsx, import_leads, parse_rows_from_file
from apps.crm.models import Lead

pytestmark = pytest.mark.django_db

LEADS_URL = "/api/v1/crm/leads/"


def _csv_file(text, name="leads.csv"):
    return SimpleUploadedFile(name, text.encode("utf-8"), content_type="text/csv")


def _xlsx_file(rows, name="leads.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile(
        name, buffer.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# --------------------------------------------------------------------------
# parse_rows_from_file()
# --------------------------------------------------------------------------


def test_parse_csv_rows():
    csv_text = "company_name,contact_name,email\nAcme,Jane,jane@acme.example\n"
    rows = parse_rows_from_file(_csv_file(csv_text))
    assert rows == [{"company_name": "Acme", "contact_name": "Jane", "email": "jane@acme.example"}]


def test_parse_csv_handles_quoted_commas():
    csv_text = 'company_name,contact_name,notes\n"Acme, Inc.",Jane,"Called, left voicemail"\n'
    rows = parse_rows_from_file(_csv_file(csv_text))
    assert rows[0]["company_name"] == "Acme, Inc."
    assert rows[0]["notes"] == "Called, left voicemail"


def test_parse_csv_tolerates_a_byte_order_mark():
    csv_text = "﻿company_name,contact_name\nAcme,Jane\n"
    rows = parse_rows_from_file(_csv_file(csv_text))
    assert rows[0]["company_name"] == "Acme"


def test_parse_csv_with_no_header_raises():
    with pytest.raises(ValueError):
        parse_rows_from_file(_csv_file(""))


def test_parse_xlsx_rows():
    rows = parse_rows_from_file(
        _xlsx_file([["company_name", "contact_name", "email"], ["Acme", "Jane", "jane@acme.example"]])
    )
    assert rows == [{"company_name": "Acme", "contact_name": "Jane", "email": "jane@acme.example"}]


def test_parse_xlsx_skips_trailing_empty_rows():
    rows = parse_rows_from_file(
        _xlsx_file([["company_name", "contact_name"], ["Acme", "Jane"], [None, None]])
    )
    assert len(rows) == 1


def test_parse_unsupported_extension_raises():
    with pytest.raises(ValueError):
        parse_rows_from_file(SimpleUploadedFile("leads.txt", b"data", content_type="text/plain"))


# --------------------------------------------------------------------------
# import_leads()
# --------------------------------------------------------------------------


def test_import_leads_creates_valid_rows(owner):
    rows = [{"company_name": "Acme", "contact_name": "Jane", "email": "jane@acme.example"}]
    summary = import_leads(rows, default_owner=owner, created_by=owner)

    assert summary["total"] == 1
    assert summary["created"] == 1
    assert summary["failed"] == 0
    lead = Lead.objects.get(pk=summary["lead_ids"][0])
    assert lead.company_name == "Acme"
    assert lead.owner_id == owner.id


def test_import_leads_reports_missing_required_fields_without_aborting_the_batch(owner):
    rows = [
        {"company_name": "", "contact_name": "Jane"},  # invalid: no company_name
        {"company_name": "Acme", "contact_name": "Bob"},  # valid
    ]
    summary = import_leads(rows, default_owner=owner)

    assert summary["total"] == 2
    assert summary["created"] == 1
    assert summary["failed"] == 1
    assert summary["errors"][0]["row"] == 1
    assert Lead.objects.filter(contact_name="Bob").exists()


def test_import_leads_normalizes_source_and_status_labels(owner):
    rows = [{"company_name": "Acme", "contact_name": "Jane", "source": "Cold Call", "status": "qualified"}]
    summary = import_leads(rows, default_owner=owner)

    lead = Lead.objects.get(pk=summary["lead_ids"][0])
    assert lead.source == Lead.Source.COLD_CALL
    assert lead.status == Lead.Status.QUALIFIED
    assert summary["warnings"] == 0


def test_import_leads_falls_back_and_warns_on_unrecognized_source(owner):
    rows = [{"company_name": "Acme", "contact_name": "Jane", "source": "Carrier Pigeon"}]
    summary = import_leads(rows, default_owner=owner)

    lead = Lead.objects.get(pk=summary["lead_ids"][0])
    assert lead.source == Lead.Source.OTHER
    assert summary["warnings"] == 1
    assert summary["failed"] == 0  # unrecognized enum value is a warning, not a failure


def test_import_leads_rejects_invalid_email_as_a_row_error(owner):
    rows = [{"company_name": "Acme", "contact_name": "Jane", "email": "not-an-email"}]
    summary = import_leads(rows, default_owner=owner)

    assert summary["failed"] == 1
    assert summary["created"] == 0


# --------------------------------------------------------------------------
# import_leads() / preview_leads() — duplicate detection (Phase 6 audit).
# Previously entirely absent from the import path: every row became a new
# Lead regardless of any existing match.
# --------------------------------------------------------------------------


def test_import_leads_skips_a_row_matching_an_existing_lead(owner):
    Lead.objects.create(company_name="Acme Inc", contact_name="Jane D", email="jane@acme.example", owner=owner)

    rows = [{"company_name": "Acme", "contact_name": "Jane", "email": "jane@acme.example"}]
    summary = import_leads(rows, default_owner=owner)

    assert summary["created"] == 0
    assert summary["failed"] == 1
    assert "Duplicate of existing lead" in summary["errors"][0]["message"]
    assert Lead.objects.filter(email="jane@acme.example").count() == 1


def test_import_leads_matches_existing_lead_case_and_format_insensitively(owner):
    Lead.objects.create(company_name="Acme Inc", contact_name="Jane D", phone="555-0100", owner=owner)

    rows = [{"company_name": "Acme", "contact_name": "Jane", "phone": "(555) 0100"}]
    summary = import_leads(rows, default_owner=owner)

    assert summary["created"] == 0
    assert summary["failed"] == 1


def test_import_leads_skips_the_second_of_two_duplicate_rows_in_the_same_file(owner):
    rows = [
        {"company_name": "Acme", "contact_name": "Jane", "email": "jane@acme.example"},
        {"company_name": "Acme Inc", "contact_name": "Jane D", "email": "jane@acme.example"},
    ]
    summary = import_leads(rows, default_owner=owner)

    assert summary["created"] == 1
    assert summary["failed"] == 1
    assert Lead.objects.filter(email="jane@acme.example").count() == 1


def test_import_leads_does_not_flag_two_blank_contact_rows_as_duplicates(owner):
    rows = [
        {"company_name": "Acme", "contact_name": "Jane"},
        {"company_name": "Beta", "contact_name": "Bob"},
    ]
    summary = import_leads(rows, default_owner=owner)

    assert summary["created"] == 2
    assert summary["failed"] == 0


def test_preview_leads_flags_a_row_matching_an_existing_lead(owner):
    from apps.crm.imports import preview_leads

    Lead.objects.create(company_name="Acme Inc", contact_name="Jane D", email="jane@acme.example", owner=owner)

    rows = [{"company_name": "Acme", "contact_name": "Jane", "email": "jane@acme.example"}]
    result = preview_leads(rows)

    assert result["valid"] == 0
    assert result["invalid"] == 1
    assert "Duplicate of existing lead" in result["errors"][0]["message"]


def test_preview_leads_flags_within_file_duplicates(owner):
    from apps.crm.imports import preview_leads

    rows = [
        {"company_name": "Acme", "contact_name": "Jane", "email": "jane@acme.example"},
        {"company_name": "Acme Inc", "contact_name": "Jane D", "email": "JANE@ACME.EXAMPLE"},
    ]
    result = preview_leads(rows)

    assert result["valid"] == 1
    assert result["invalid"] == 1
    assert "same file" in result["errors"][0]["message"]


# --------------------------------------------------------------------------
# export_leads_csv() / export_leads_xlsx()
# --------------------------------------------------------------------------


def test_export_leads_csv_contains_the_row(owner):
    Lead.objects.create(company_name="Acme", contact_name="Jane", email="jane@acme.example", owner=owner)
    content = export_leads_csv(Lead.objects.all()).decode("utf-8")
    assert "Acme" in content
    assert "Jane" in content
    assert owner.email in content


def test_export_leads_xlsx_contains_the_row(owner):
    Lead.objects.create(company_name="Acme", contact_name="Jane", owner=owner)
    content = export_leads_xlsx(Lead.objects.all())
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    values = [cell.value for row in workbook.active.iter_rows() for cell in row]
    assert "Acme" in values
    assert "Jane" in values


# --------------------------------------------------------------------------
# CSV/XLSX formula-injection guard (CWE-1236) — a company_name/contact_name/
# notes value starting with a spreadsheet formula trigger character must be
# neutralized in the exported file, since these are free-text fields a
# caller (including an attacker via the Google Sheets import path) controls
# and this file is opened directly in Excel/Sheets by an employee.
# --------------------------------------------------------------------------


@pytest.mark.parametrize("trigger", ["=", "+", "-", "@", "\t", "\r"])
def test_export_leads_csv_neutralizes_formula_trigger_characters(owner, trigger):
    payload = f'{trigger}cmd|\'/c calc\'!A1'
    Lead.objects.create(company_name=payload, contact_name="Jane", owner=owner)
    rows = list(csv.reader(io.StringIO(export_leads_csv(Lead.objects.all()).decode("utf-8"))))
    company_name_cell = rows[1][EXPORT_COLUMNS.index("company_name")]
    # The neutralized (leading-quote-prefixed) form is the actual cell
    # value — never the raw payload as its own cell, which a spreadsheet
    # app would treat as a formula the moment the file is opened.
    assert company_name_cell == f"'{payload}"
    assert company_name_cell.startswith("'")


@pytest.mark.parametrize("trigger", ["=", "+", "-", "@"])
def test_export_leads_xlsx_neutralizes_formula_trigger_characters(owner, trigger):
    payload = f"{trigger}HYPERLINK(\"http://evil.example\",\"click\")"
    Lead.objects.create(company_name=payload, contact_name="Jane", owner=owner)
    content = export_leads_xlsx(Lead.objects.all())
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    values = [cell.value for row in workbook.active.iter_rows() for cell in row]
    assert payload not in values
    assert f"'{payload}" in values


def test_export_leads_csv_leaves_ordinary_values_untouched(owner):
    Lead.objects.create(company_name="Acme Corp", contact_name="Jane", owner=owner)
    content = export_leads_csv(Lead.objects.all()).decode("utf-8")
    assert "Acme Corp" in content
    assert "'Acme Corp" not in content


# --------------------------------------------------------------------------
# API actions
# --------------------------------------------------------------------------


def test_import_leads_via_api(api_client, owner):
    api_client.force_authenticate(owner)
    csv_text = "company_name,contact_name,email\nAcme,Jane,jane@acme.example\n"

    response = api_client.post(f"{LEADS_URL}import/", {"file": _csv_file(csv_text)}, format="multipart")

    assert response.status_code == 200
    assert response.data["created"] == 1
    assert Lead.objects.filter(company_name="Acme", owner=owner).exists()


def test_import_leads_via_api_requires_a_file(api_client, owner):
    api_client.force_authenticate(owner)
    response = api_client.post(f"{LEADS_URL}import/", {}, format="multipart")
    assert response.status_code == 400


def test_import_leads_via_api_rejects_unauthenticated(api_client):
    response = api_client.post(f"{LEADS_URL}import/", {"file": _csv_file("company_name,contact_name\n")}, format="multipart")
    assert response.status_code == 401


def test_export_leads_via_api_csv(api_client, owner):
    Lead.objects.create(company_name="Acme", contact_name="Jane", owner=owner)
    api_client.force_authenticate(owner)

    response = api_client.get(f"{LEADS_URL}export/")

    assert response.status_code == 200
    assert response["Content-Type"] == "text/csv"
    assert b"Acme" in response.content


def test_export_leads_via_api_xlsx(api_client, owner):
    Lead.objects.create(company_name="Acme", contact_name="Jane", owner=owner)
    api_client.force_authenticate(owner)

    response = api_client.get(f"{LEADS_URL}export/", {"export_format": "xlsx"})

    assert response.status_code == 200
    assert response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_export_leads_via_api_only_includes_leads_the_requester_can_see(api_client, owner, django_user_model):
    other_owner = django_user_model.objects.create_user(email="other-export-owner@example.com", password="x")
    Lead.objects.create(company_name="Mine", contact_name="Jane", owner=owner)
    Lead.objects.create(company_name="NotMine", contact_name="Bob", owner=other_owner)
    api_client.force_authenticate(owner)

    response = api_client.get(f"{LEADS_URL}export/")

    assert b"Mine" in response.content
    assert b"NotMine" not in response.content


def test_export_leads_via_api_rejects_invalid_format(api_client, owner):
    api_client.force_authenticate(owner)
    response = api_client.get(f"{LEADS_URL}export/", {"export_format": "pdf"})
    assert response.status_code == 400
