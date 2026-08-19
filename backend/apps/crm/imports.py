"""Final-completion-pass: real CSV/XLSX bulk import and export for Leads.

Import is a per-row, best-effort operation: one malformed row does not
abort the whole file — each row is created in its own savepoint
(``transaction.atomic()`` nested inside the caller's own transaction), and
a row that fails validation is recorded in the returned summary rather
than raised. This matches how a human actually uses a bulk-import feature:
upload a spreadsheet of imperfect data, fix only the rows that were
rejected, re-upload just those.

Export streams the SAME rows a client would see via the ordinary list
endpoint (whatever queryset the caller passes in — already ownership- and
filter-scoped by the view) — never a hardcoded/unscoped dump.
"""
import csv
import io

import openpyxl
from django.core.exceptions import ValidationError
from django.db import transaction

from .models import Lead

#: Column headers accepted on import, and the model field each maps to.
#: "company_name"/"contact_name" are required; everything else is optional.
IMPORT_COLUMNS = ["company_name", "contact_name", "email", "phone", "source", "status", "notes"]

#: Columns written on export, in order — the header row a re-import of an
#: exported file would itself need to match.
EXPORT_COLUMNS = [
    "id", "company_name", "contact_name", "email", "phone", "source", "status",
    "owner", "notes", "created_at", "updated_at",
]


def _normalize_choice(raw_value, choices_class, default):
    """Best-effort match of a free-text import cell against a
    ``TextChoices`` class: exact value match, then case-insensitive label
    match (e.g. "Cold Call" -> "COLD_CALL"), falling back to ``default``
    rather than rejecting the row outright — an unrecognized source/status
    is a data-quality issue worth flagging for review, not a hard failure
    that blocks the rest of a large import.
    """
    if not raw_value:
        return default, False
    raw_value = raw_value.strip()
    values = {choice.value: choice.value for choice in choices_class}
    if raw_value.upper() in values:
        return raw_value.upper(), False
    labels = {str(choice.label).lower(): choice.value for choice in choices_class}
    if raw_value.lower() in labels:
        return labels[raw_value.lower()], False
    return default, True


def parse_rows_from_file(uploaded_file):
    """Parse ``uploaded_file`` (a Django ``UploadedFile``) into a list of
    plain dicts, one per data row. Format is detected from the filename
    extension (``.csv`` vs ``.xlsx``) — the one piece of metadata that's
    always present and never ambiguous, unlike ``content_type`` (browsers
    are inconsistent about what they send for CSV).

    Raises ``ValueError`` for an unsupported extension or a file with no
    header row — both are caller/file-format problems, not per-row data
    problems, so they abort the whole import rather than being folded
    into the per-row error list.
    """
    name = (uploaded_file.name or "").lower()

    if name.endswith(".xlsx"):
        workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = [str(cell).strip() if cell is not None else "" for cell in next(rows_iter)]
        except StopIteration:
            raise ValueError("The uploaded file has no header row.")
        rows = []
        for raw_row in rows_iter:
            if raw_row is None or all(cell is None for cell in raw_row):
                continue  # openpyxl can yield trailing fully-empty rows
            row = {header[i]: ("" if raw_row[i] is None else str(raw_row[i]).strip()) for i in range(len(header)) if i < len(raw_row)}
            rows.append(row)
        return rows

    if name.endswith(".csv"):
        text = uploaded_file.read().decode("utf-8-sig")  # -sig: tolerate a BOM from Excel-exported CSVs
        reader = csv.DictReader(io.StringIO(text))
        if reader.fieldnames is None:
            raise ValueError("The uploaded file has no header row.")
        return [{key: (value or "").strip() for key, value in row.items() if key} for row in reader]

    raise ValueError("Unsupported file type — upload a .csv or .xlsx file.")


@transaction.atomic
def import_leads(rows, *, default_owner=None, created_by=None):
    """Create a ``Lead`` for each row in ``rows`` (as produced by
    ``parse_rows_from_file()``). Returns a summary dict:

    ``{"total": int, "created": int, "failed": int, "warnings": int,
    "errors": [{"row": int, "message": str}], "lead_ids": [int, ...]}``

    ``row`` in each error is 1-indexed against the DATA rows (excluding
    the header), matching what a spreadsheet user would call "row 1".
    An unrecognized source/status value doesn't fail the row — it falls
    back to the model's own default and is counted in ``warnings``
    instead, since the row is still genuinely creatable.
    """
    errors = []
    warnings_count = 0
    lead_ids = []

    for index, row in enumerate(rows, start=1):
        company_name = row.get("company_name", "").strip()
        contact_name = row.get("contact_name", "").strip()

        if not company_name or not contact_name:
            errors.append({"row": index, "message": "company_name and contact_name are both required."})
            continue

        source, source_unrecognized = _normalize_choice(row.get("source", ""), Lead.Source, Lead.Source.OTHER)
        status, status_unrecognized = _normalize_choice(row.get("status", ""), Lead.Status, Lead.Status.NEW)
        if source_unrecognized or status_unrecognized:
            warnings_count += 1

        email = row.get("email", "").strip()
        try:
            with transaction.atomic():
                lead = Lead(
                    company_name=company_name,
                    contact_name=contact_name,
                    email=email,
                    phone=row.get("phone", "").strip(),
                    source=source,
                    status=status,
                    notes=row.get("notes", "").strip(),
                    owner=default_owner,
                    created_by=created_by,
                    updated_by=created_by,
                )
                lead.full_clean(exclude=["converted_customer"])
                lead.save()
        except ValidationError as exc:
            errors.append({"row": index, "message": "; ".join(f"{field}: {', '.join(msgs)}" for field, msgs in exc.message_dict.items())})
            continue

        lead_ids.append(lead.id)

    return {
        "total": len(rows),
        "created": len(lead_ids),
        "failed": len(errors),
        "warnings": warnings_count,
        "errors": errors,
        "lead_ids": lead_ids,
    }


#: How many normalized rows a preview echoes back. A preview is meant to
#: let a human eyeball what they're about to import, not to stream an
#: entire spreadsheet through a JSON response.
PREVIEW_SAMPLE_SIZE = 20


def preview_leads(rows, *, sample_size=PREVIEW_SAMPLE_SIZE):
    """Staff-management pass: the PREVIEW stage of
    source -> adapter -> validation -> preview -> confirm -> CRM records.

    Runs the exact same per-row validation ``import_leads()`` does — the
    required-field check, the ``_normalize_choice()`` fallbacks, and Django's
    own ``full_clean()`` — but creates NOTHING. Read-only by construction:
    every candidate ``Lead`` is built unsaved and validated in memory, so
    there is no transaction to roll back and no id to leak.

    Returns::

        {"total": int, "valid": int, "invalid": int, "warnings": int,
         "errors": [{"row": int, "message": str}],
         "sample": [{...normalized row...}, ...]}

    ``sample`` holds the first ``sample_size`` rows AS THEY WOULD BE
    CREATED (post-normalization), which is what makes the preview useful:
    a user sees that "Cold Call" became ``COLD_CALL`` before confirming.
    """
    errors = []
    warnings_count = 0
    valid_count = 0
    sample = []

    for index, row in enumerate(rows, start=1):
        company_name = row.get("company_name", "").strip()
        contact_name = row.get("contact_name", "").strip()

        if not company_name or not contact_name:
            errors.append({"row": index, "message": "company_name and contact_name are both required."})
            continue

        source, source_unrecognized = _normalize_choice(row.get("source", ""), Lead.Source, Lead.Source.OTHER)
        status, status_unrecognized = _normalize_choice(row.get("status", ""), Lead.Status, Lead.Status.NEW)
        if source_unrecognized or status_unrecognized:
            warnings_count += 1

        candidate = {
            "company_name": company_name,
            "contact_name": contact_name,
            "email": row.get("email", "").strip(),
            "phone": row.get("phone", "").strip(),
            "source": source,
            "status": status,
            "notes": row.get("notes", "").strip(),
        }

        try:
            Lead(**candidate).full_clean(exclude=["converted_customer", "owner"])
        except ValidationError as exc:
            errors.append(
                {
                    "row": index,
                    "message": "; ".join(
                        f"{field}: {', '.join(msgs)}" for field, msgs in exc.message_dict.items()
                    ),
                }
            )
            continue

        valid_count += 1
        if len(sample) < sample_size:
            sample.append({"row": index, **candidate})

    return {
        "total": len(rows),
        "valid": valid_count,
        "invalid": len(errors),
        "warnings": warnings_count,
        "errors": errors,
        "sample": sample,
    }


#: Export columns that carry the LEAD's own contact PII. Omitted
#: entirely (header row included) from an employee-facing export — see
#: ``export_leads_csv()``'s ``include_pii`` argument.
PII_EXPORT_COLUMNS = ("email", "phone")


#: Export columns only a SUPER ADMIN may receive. "source" (Lead Source)
#: is hidden from BOTH Manager and Employee — a stricter, separate rule
#: from PII_EXPORT_COLUMNS above (which a Manager still sees). See
#: ``apps.core.serializers.SuperAdminOnlyFieldsMixin``.
SUPER_ADMIN_ONLY_EXPORT_COLUMNS = ("source",)


def _export_columns(include_pii, include_source=True):
    columns = list(EXPORT_COLUMNS)
    if not include_pii:
        columns = [column for column in columns if column not in PII_EXPORT_COLUMNS]
    if not include_source:
        columns = [column for column in columns if column not in SUPER_ADMIN_ONLY_EXPORT_COLUMNS]
    return columns


def export_leads_csv(queryset, *, include_pii=True, include_source=True):
    """Return CSV file bytes for ``queryset`` (already ownership/filter
    scoped by the caller).

    ``include_pii=False`` drops the lead ``email``/``phone`` columns —
    both the header and every value — so a downloadable file can never
    become the side channel that hands an Employee the raw contact
    details the API itself refuses to return. The caller (see
    ``apps.crm.views.LeadViewSet.export_leads_action()``) decides, from
    the requesting user's role; the default stays ``True`` so every
    existing non-HTTP caller keeps its current behavior.

    ``include_source=False`` likewise drops the ``source`` (Lead Source)
    column for any reader below Super Admin — the same reasoning applied
    to a stricter field. Defaults stay ``True`` for both so every existing
    non-HTTP caller keeps its current behavior.
    """
    columns = _export_columns(include_pii, include_source)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(columns)
    for lead in queryset.select_related("owner"):
        writer.writerow(_export_row(lead, include_pii=include_pii, include_source=include_source))
    return buffer.getvalue().encode("utf-8")


def export_leads_xlsx(queryset, *, include_pii=True, include_source=True):
    """Return XLSX file bytes for ``queryset`` (already ownership/filter
    scoped by the caller) — see ``export_leads_csv()`` for ``include_pii``
    and ``include_source``.
    """
    columns = _export_columns(include_pii, include_source)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Leads"
    sheet.append(columns)
    for lead in queryset.select_related("owner"):
        sheet.append(_export_row(lead, include_pii=include_pii, include_source=include_source))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _export_row(lead, *, include_pii=True, include_source=True):
    values = {
        "id": lead.id,
        "company_name": lead.company_name,
        "contact_name": lead.contact_name,
        "email": lead.email,
        "phone": lead.phone,
        "source": lead.source,
        "status": lead.status,
        "owner": lead.owner.email if lead.owner_id else "",
        "notes": lead.notes,
        "created_at": lead.created_at.isoformat(),
        "updated_at": lead.updated_at.isoformat(),
    }
    return [values[column] for column in _export_columns(include_pii, include_source)]
